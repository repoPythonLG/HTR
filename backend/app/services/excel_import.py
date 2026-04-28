from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.services.normalization import normalize_field


HEADER_ALIASES = {
    # Legacy/previous templates
    "employee_id": "employee_id",
    "employee id": "employee_id",
    "employee code": "employee_id",
    "emp id": "employee_id",
    "employee_name": "employee_name",
    "employee name": "employee_name",
    "employee": "employee_name",
    "name": "employee_name",
    "department": "department",
    "dept": "department",
    "start_date": "start_date",
    "start date": "start_date",
    "trip_start": "start_date",
    "end_date": "end_date",
    "end date": "end_date",
    "trip_end": "end_date",
    "destination_city": "destination_city",
    "destination city": "destination_city",
    "destination": "destination_city",
    "city": "destination_city",
    "claim_total": "claim_total",
    "claim total": "claim_total",
    "total": "claim_total",
    "currency": "currency",
    "hotel_rate": "hotel_rate",
    "hotel rate": "hotel_rate",
    "room_rate": "hotel_rate",
    "room rate": "hotel_rate",
    "nightly_rate": "hotel_rate",
    "nightly rate": "hotel_rate",
    "nights": "nights",
    "receipt_number": "receipt_number",
    "receipt number": "receipt_number",
    "receipt_no": "receipt_number",
    "receipt": "receipt_number",
    "merchant": "merchant",
    "vendor": "merchant",
    "booking_channel": "booking_channel",
    "booking channel": "booking_channel",
    "meal_included": "meal_included",
    "meal included": "meal_included",
    "flight_class": "flight_class",
    "flight class": "flight_class",

    # SABIC travel register schema
    "trip number": "trip_number",
    "trip_number": "trip_number",
    "from region": "from_region",
    "from country": "from_country",
    "from city": "from_city",
    "to region": "to_region",
    "to country": "to_country",
    "to city": "to_city",
    "trip activity": "trip_activity",
    "trip start date": "start_date",
    "trip end date": "end_date",
    "trip boundary": "trip_boundary",
    "trip settlement date": "trip_settlement_date",
    "expense type": "expense_type",
    "masked id": "masked_id",
    "trip duration": "trip_duration",
    "expense amount in sar": "claim_total",
    "expense amount": "claim_total",
}


def _clean_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _to_iso_date(value: Any) -> Optional[str]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    for fmt in [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d-%b-%Y",
        "%d-%B-%Y",
    ]:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def _to_int(value: Any, default: int = 0) -> int:
    return int(round(_to_float(value, float(default))))


def _to_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "included"}


def _normalize_row(raw_row: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {}
    for key, value in raw_row.items():
        key_clean = _clean_header(key)
        canonical = HEADER_ALIASES.get(key_clean)
        if canonical:
            normalized[canonical] = value
    return normalized


def _derive_duration(start_date: Optional[str], end_date: Optional[str], raw_duration: Any) -> Optional[int]:
    if raw_duration is not None and str(raw_duration).strip() != "":
        value = _to_int(raw_duration, 0)
        if value > 0:
            return value

    if not start_date or not end_date:
        return None

    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        return None

    return max(1, (end - start).days + 1)


def _build_claim_payload(row: Dict[str, Any], row_number: int, source_name: str) -> Dict[str, Any]:
    masked_id = str(row.get("masked_id") or "").strip()
    employee_id = str(row.get("employee_id") or masked_id or f"UNMAPPED-{row_number}").strip()
    employee_name = str(row.get("employee_name") or (f"Employee {masked_id}" if masked_id else "Unknown Employee")).strip()

    start_date = _to_iso_date(row.get("start_date"))
    end_date = _to_iso_date(row.get("end_date"))
    trip_settlement_date = _to_iso_date(row.get("trip_settlement_date"))

    to_city = str(row.get("to_city") or row.get("destination_city") or "").strip().title() or None
    destination_city = to_city

    claim_total = _to_float(row.get("claim_total"), 0.0)
    hotel_rate = _to_float(row.get("hotel_rate"), 0.0)
    nights = max(1, _to_int(row.get("nights"), 1))
    if claim_total <= 0 and hotel_rate > 0:
        claim_total = hotel_rate * nights

    trip_duration_days = _derive_duration(start_date, end_date, row.get("trip_duration"))

    currency = str(row.get("currency") or "SAR").strip().upper() or "SAR"

    trip_number = str(row.get("trip_number") or "").strip() or None
    from_region = str(row.get("from_region") or "").strip() or None
    from_country = str(row.get("from_country") or "").strip() or None
    from_city = str(row.get("from_city") or "").strip().title() or None
    to_region = str(row.get("to_region") or "").strip() or None
    to_country = str(row.get("to_country") or "").strip() or None
    trip_activity = str(row.get("trip_activity") or "").strip() or None
    trip_boundary = str(row.get("trip_boundary") or "").strip().title() or None
    expense_type = str(row.get("expense_type") or "").strip() or None

    department = str(row.get("department") or trip_activity or "Travel").strip()

    receipt_number = str(row.get("receipt_number") or f"AUTO-{row_number:04d}").strip()
    merchant = str(row.get("merchant") or "Travel Vendor").strip()
    booking_channel = str(row.get("booking_channel") or "compliant").strip().lower().replace(" ", "_")
    meal_included = _to_bool(row.get("meal_included"))
    flight_class = str(row.get("flight_class") or "economy class").strip().lower()

    evidence_lines = [
        f"Trip Number: {trip_number or '-'}",
        f"Masked ID: {masked_id or '-'}",
        f"From: {from_city or '-'}, {from_country or '-'} ({from_region or '-'})",
        f"To: {to_city or '-'}, {to_country or '-'} ({to_region or '-'})",
        f"Trip Activity: {trip_activity or '-'}",
        f"Trip Boundary: {trip_boundary or '-'}",
        f"Trip Start: {start_date or '-'}",
        f"Trip End: {end_date or '-'}",
        f"Trip Settlement: {trip_settlement_date or '-'}",
        f"Trip Duration Days: {trip_duration_days or '-'}",
        f"Expense Type: {expense_type or '-'}",
        f"Expense Amount in SAR: {claim_total:.2f}",
        f"Entry no: {receipt_number}",
        f"Merchant: {merchant}",
        f"Booking channel: {booking_channel}",
        f"Flight class: {flight_class}",
    ]
    if meal_included:
        evidence_lines.append("Breakfast included")

    extracted_fields = [
        {"name": "trip_number", "value": trip_number or "", "confidence": 0.99},
        {"name": "masked_id", "value": masked_id, "confidence": 0.99},
        {"name": "from_region", "value": from_region or "", "confidence": 0.98},
        {"name": "from_country", "value": from_country or "", "confidence": 0.98},
        {"name": "from_city", "value": from_city or "", "confidence": 0.98},
        {"name": "to_region", "value": to_region or "", "confidence": 0.98},
        {"name": "to_country", "value": to_country or "", "confidence": 0.98},
        {"name": "to_city", "value": to_city or "", "confidence": 0.98},
        {"name": "trip_activity", "value": trip_activity or "", "confidence": 0.97},
        {"name": "trip_boundary", "value": trip_boundary or "", "confidence": 0.97},
        {"name": "trip_settlement_date", "value": trip_settlement_date or "", "confidence": 0.96},
        {"name": "expense_type", "value": expense_type or "", "confidence": 0.97},
        {"name": "trip_duration", "value": str(trip_duration_days or ""), "confidence": 0.96},
        {"name": "total", "value": f"{claim_total:.2f}", "confidence": 0.99},
        {"name": "city", "value": destination_city or "", "confidence": 0.98},
        {"name": "merchant", "value": merchant, "confidence": 0.95},
        {"name": "receipt_number", "value": receipt_number, "confidence": 0.95},
        {"name": "booking_channel", "value": booking_channel, "confidence": 0.9},
        {"name": "meal_included", "value": "true" if meal_included else "false", "confidence": 0.9},
        {"name": "flight_class", "value": flight_class, "confidence": 0.9},
    ]

    for field in extracted_fields:
        field["normalized_value"] = normalize_field(field["name"], str(field["value"]))

    return {
        "claim": {
            "employee_id": employee_id,
            "employee_name": employee_name,
            "department": department,
            "start_date": start_date,
            "end_date": end_date,
            "trip_settlement_date": trip_settlement_date,
            "trip_number": trip_number,
            "from_region": from_region,
            "from_country": from_country,
            "from_city": from_city,
            "to_region": to_region,
            "to_country": to_country,
            "to_city": to_city,
            "trip_activity": trip_activity,
            "trip_boundary": trip_boundary,
            "expense_type": expense_type,
            "masked_id": masked_id or None,
            "trip_duration_days": trip_duration_days,
            "destination_city": destination_city,
            "claim_total": claim_total,
            "currency": currency,
            "source_reference": f"{source_name}:row:{row_number}",
        },
        "document": {
            "file_name": f"{source_name}-row-{row_number}.txt",
            "text": "\n".join(evidence_lines),
            "document_type": "travel_expense_register",
            "mime_type": "application/x-structured-claim",
            "fields": extracted_fields,
        },
    }


def _build_raw_row(headers: List[str], values: List[Any]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for index in range(len(headers)):
        row[headers[index]] = values[index] if index < len(values) else ""
    return row


def _validate_row(normalized: Dict[str, Any], row_number: int) -> Optional[str]:
    amount = _to_float(normalized.get("claim_total"), 0.0)
    if amount <= 0 and _to_float(normalized.get("hotel_rate"), 0.0) <= 0:
        return f"Row {row_number}: missing valid expense amount"

    if not _to_iso_date(normalized.get("start_date")) and not _to_iso_date(normalized.get("end_date")):
        return f"Row {row_number}: missing trip start/end date"

    return None


def parse_claim_rows(file_name: str, content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    records: List[Dict[str, Any]] = []

    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for idx, raw_row in enumerate(reader, start=2):
            normalized = _normalize_row(raw_row)
            if not normalized:
                errors.append(f"Row {idx}: no recognized columns")
                continue

            validation_error = _validate_row(normalized, idx)
            if validation_error:
                errors.append(validation_error)
                continue

            records.append(_build_claim_payload(normalized, idx, file_name))
        return records, errors

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active

    headers: List[str] = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [cell if cell is not None else "" for cell in row]
        if row_idx == 1:
            headers = [str(cell).strip() for cell in values]
            continue

        if not any(str(cell).strip() for cell in values):
            continue

        raw_row = _build_raw_row(headers, values)
        normalized = _normalize_row(raw_row)
        if not normalized:
            errors.append(f"Row {row_idx}: no recognized columns")
            continue

        validation_error = _validate_row(normalized, row_idx)
        if validation_error:
            errors.append(validation_error)
            continue

        records.append(_build_claim_payload(normalized, row_idx, file_name))

    return records, errors


def parse_spreadsheet_preview(file_name: str, content: bytes, max_rows: int = 2000) -> Tuple[List[str], List[List[str]]]:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], []
        headers = [str(item).strip() for item in rows[0]]
        data_rows: List[List[str]] = []
        for row in rows[1 : max_rows + 1]:
            values = [str(cell).strip() for cell in row]
            if any(values):
                data_rows.append(values)
        return headers, data_rows

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    data: List[List[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values: List[str] = []
        for cell in row:
            if cell is None:
                values.append("")
            elif isinstance(cell, datetime):
                values.append(cell.date().isoformat())
            elif isinstance(cell, date):
                values.append(cell.isoformat())
            else:
                values.append(str(cell).strip())
        if any(values):
            data.append(values)
        if len(data) >= max_rows + 1:
            break

    if not data:
        return [], []
    headers = data[0]
    return headers, data[1:]
